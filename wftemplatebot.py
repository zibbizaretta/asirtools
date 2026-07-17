import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
import re
import io
import copy

# --- 1. HAFIZA (SESSION STATE) ---
if 'user_prefs' not in st.session_state:
    st.session_state['user_prefs'] = {}

# --- 2. YARDIMCI VE LOJİSTİK FONKSİYONLAR ---

def get_dim_val(pattern, text):
    match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
    if match:
        try:
            return float(match.group(1).replace(',', '.'))
        except:
            return None
    return None

def extract_overall_dims(text):
    if pd.isna(text):
        return None, None, None
    text = str(text)
    
    dia = get_dim_val(r'(?:Diameter|Çap|Dia|Ø)\s*[:\-\s]\s*(\d+(?:[.,]\d+)?)', text)
    w = get_dim_val(r'(?:Width|Genişlik|Side to Side|\bW\b)\s*[:\-\s]\s*(\d+(?:[.,]\d+)?)', text)
    h = get_dim_val(r'(?:Height|Yükseklik|Top to Bottom|\bH\b)\s*[:\-\s]\s*(\d+(?:[.,]\d+)?)', text)
    d = get_dim_val(r'(?:Depth|Derinlik|Front to Back|\bD\b)\s*[:\-\s]\s*(\d+(?:[.,]\d+)?)', text)
    
    if dia:
        if not w: w = dia
        if not d: d = dia
            
    if not any([w, h, d]):
        size_match = re.search(
            r'(?:Size|Ölçü|Dimension|Boyut)?[:\s]*(\d+(?:[.,]\d+)?)\s*[xX×]\s*(\d+(?:[.,]\d+)?)(?:\s*[xX×]\s*(\d+(?:[.,]\d+)?))?', 
            text
        )
        if size_match:
            h = float(size_match.group(1).replace(',', '.'))
            w = float(size_match.group(2).replace(',', '.'))
            if size_match.group(3): 
                d = float(size_match.group(3).replace(',', '.'))
                
    return h, w, d

def convert_to_inch(val):
    if val: 
        return round(val * 0.393701, 2)
    return None

def translate_features(text, do_conversion):
    if pd.isna(text): 
        return ""
    text = str(text)
    text = re.sub(r'Ø\s*[:\-]?\s*', 'Diameter: ', text)
    text = re.sub(r'(?i)\bÇap\s*[:\-]?\s*', 'Diameter: ', text)
    
    if not do_conversion: 
        return text

    def c_in(m): return f"{round(float(m.group(1).replace(',', '.')) * 0.393701, 2)}"
    def c_mm(m): return f"{round(float(m.group(1).replace(',', '.')) * 0.0393701, 2)}"
    def c_kg(m): return f"{round(float(m.group(1).replace(',', '.')) * 2.20462, 2)}"
    def c_ml(m): return f"{round(float(m.group(1).replace(',', '.')) * 0.033814, 2)}"

    text = re.sub(r'(?<![a-zA-Z0-9])(\d+(?:[\.,]\d+)?)\s*cm(?![a-zA-Z])', lambda m: c_in(m) + ' inches', text, flags=re.IGNORECASE)
    text = re.sub(r'(?<![a-zA-Z0-9])(\d+(?:[\.,]\d+)?)\s*mm(?![a-zA-Z])', lambda m: c_mm(m) + ' inches', text, flags=re.IGNORECASE)
    text = re.sub(r'(?<![a-zA-Z0-9])(\d+(?:[\.,]\d+)?)\s*kg(?![a-zA-Z])', lambda m: c_kg(m) + ' lbs', text, flags=re.IGNORECASE)
    text = re.sub(r'(?<![a-zA-Z0-9])(\d+(?:[\.,]\d+)?)\s*ml(?![a-zA-Z])', lambda m: c_ml(m) + ' fl oz', text, flags=re.IGNORECASE)
    text = re.sub(r'(?<![a-zA-Z0-9:])(\d+(?:[\.,]\d+)?)(?=\s*[xX×]\s*\d)', c_in, text)
    text = re.sub(r'(?<![a-zA-Z])\bcm\b(?![a-zA-Z])', 'inches', text, flags=re.IGNORECASE)
    text = re.sub(r'(?<![a-zA-Z])\bmm\b(?![a-zA-Z])', 'inches', text, flags=re.IGNORECASE)
    
    return text

def calculate_freight_class_total(total_weight_lbs, total_volume_in3):
    vol_ft3 = total_volume_in3 / 1728
    if vol_ft3 == 0: 
        return "60"
    dens = total_weight_lbs / vol_ft3
    if dens < 1: return "400"
    elif dens < 2: return "300"
    elif dens < 4: return "200"
    elif dens < 6: return "150"
    elif dens < 8: return "125"
    elif dens < 10: return "100"
    else: return "60"

def extract_bedding_info(features, description, raw_h, raw_w):
    text_lower = (str(features) + " " + str(description)).lower()
    total_pieces = 0
    pillow_match = re.search(r'(?:pillowcase|pillow case|yastık kılıfı).*?(\d+)\s*(?:piece|adet|pcs)', text_lower)
    
    if pillow_match: total_pieces += int(pillow_match.group(1))
    elif 'pillowcase' in text_lower or 'yastık kılıfı' in text_lower: total_pieces += 1
        
    if any(x in text_lower for x in ['duvet cover', 'quilt cover', 'nevresim']): total_pieces += 1
    if any(x in text_lower for x in ['fitted sheet', 'çarşaf']): total_pieces += 1
    if any(x in text_lower for x in ['flat sheet', 'düz çarşaf']): total_pieces += 1
    if any(x in text_lower for x in ['bedspread', 'yatak örtüsü']): total_pieces += 1
    if any(x in text_lower for x in ['comforter', 'yorgan']): total_pieces += 1
    if any(x in text_lower for x in ['blanket', 'battaniye']): total_pieces += 1
        
    pieces_str = str(total_pieces) if total_pieces > 0 else ""
    
    set_single = ""
    if total_pieces > 1: set_single = "Set (matching pieces included)"
    elif total_pieces == 1: set_single = "Single Piece"
        
    material = ""
    if 'cotton blend' in text_lower or ('cotton' in text_lower and 'polyester' in text_lower): material = "Cotton Blend"
    elif 'cotton' in text_lower or 'pamuk' in text_lower: material = "Cotton"
    elif 'microfiber' in text_lower or 'mikrofiber' in text_lower: material = "Microfiber"
    elif 'polyester' in text_lower: material = "Polyester"
    elif 'satin' in text_lower or 'saten' in text_lower: material = "Satin"
    elif 'linen' in text_lower or 'keten' in text_lower: material = "Linen"
    elif 'flannel' in text_lower or 'pazen' in text_lower: material = "Flannel"
    elif 'silk' in text_lower or 'ipek' in text_lower: material = "Silk"
    elif 'velvet' in text_lower or 'kadife' in text_lower: material = "Velour"
    elif 'rayon' in text_lower or 'viskon' in text_lower: material = "Rayon"
    
    prod_type = ""
    is_bedding = False
    
    if 'duvet cover' in text_lower or 'nevresim' in text_lower or 'quilt cover' in text_lower: 
        prod_type = "Duvet Cover"; is_bedding = True
    elif 'bedspread' in text_lower or 'yatak örtüsü' in text_lower: 
        prod_type = "Bedspread"; is_bedding = True
    elif 'quilt' in text_lower: 
        prod_type = "Quilt"; is_bedding = True
    elif 'comforter' in text_lower or 'yorgan' in text_lower: 
        prod_type = "Comforter"; is_bedding = True
    elif 'coverlet' in text_lower: 
        prod_type = "Coverlet"; is_bedding = True
    elif 'pillowcase' in text_lower or 'yastık kılıfı' in text_lower or 'sham' in text_lower: 
        prod_type = "Sham"
        
    bed_size = ""
    max_dim = max(float(raw_h or 0), float(raw_w or 0))
    
    if max_dim > 0 and (is_bedding or prod_type == "Sham"):
        if max_dim >= 250: bed_size = "California King"
        elif max_dim >= 230: bed_size = "King"
        elif max_dim >= 200: bed_size = "Queen"
        elif max_dim >= 180: bed_size = "Full / Double"
        else: bed_size = "Twin"
        
    new_name = str(description)
    new_name = re.sub(r'\s*\([A-Z]{2,3}\)\s*', ' ', new_name)
    
    size_words = [r'\bSingle XXL\b', r'\bSingle XL\b', r'\bSingle\b', r'\bDouble\b', r'\bKing\b', r'\bSuper King\b', r'\bMega King\b', r'\bQueen\b', r'\bTwin\b', r'\bFull\b']
    for word in size_words: 
        new_name = re.sub(word, '', new_name, flags=re.IGNORECASE)
        
    new_name = re.sub(r'\s+', ' ', new_name).strip()
    if bed_size and prod_type != "Sham" and is_bedding: 
        new_name = f"{bed_size} {new_name}"
        
    return {'pieces': pieces_str, 'set_single': set_single, 'material': material, 'prod_type': prod_type, 'bed_size': bed_size, 'new_name': new_name}

def generate_bedding_note(text, h_val, w_val, bed_size, is_us):
    if not is_us or not text: return ""
    text_lower = str(text).lower()
    is_bedding = any(k in text_lower for k in ['duvet', 'quilt', 'bedspread', 'nevresim', 'yorgan', 'yatak örtüsü', 'comforter', 'coverlet'])
    is_pillow = any(k in text_lower for k in ['pillowcase', 'yastık kılıfı', 'sham'])

    if not (is_bedding or is_pillow): return ""

    h_in = round(float(h_val) * 0.393701, 2) if h_val else ""
    w_in = round(float(w_val) * 0.393701, 2) if w_val else ""
    dims_str = f"{h_in} x {w_in}" if h_in and w_in else f"{h_in or w_in}"
    
    if is_bedding and bed_size: 
        return f"This beautiful set is crafted to European standards. Measuring {dims_str} inches, it beautifully complements your {bed_size} bed, offering a cozy and elegant drape."
    elif is_pillow: 
        return f"Crafted in Türkiye, these pillowcases measure {dims_str} inches. They are a wonderful fit for standard US pillows, bringing a touch of European comfort to your bedroom."
    return ""

def get_brand_by_category(category_text):
    if pd.isna(category_text) or not str(category_text).strip(): return ""
    cat_lower = str(category_text).lower()
    if "sofa" in cat_lower or "koltuk" in cat_lower or "kanepe" in cat_lower: return "Hanah Home"
    elif "wall deco" in cat_lower or "duvar" in cat_lower or "tablo" in cat_lower: return "Wallity"
    elif "rug" in cat_lower or "carpet" in cat_lower or "halı" in cat_lower or "kilim" in cat_lower: return "Conceptum Hypnose"
    elif "kitchen" in cat_lower or "mutfak" in cat_lower: return "Hermia Concept"
    elif "lighting" in cat_lower or "aydınlatma" in cat_lower or "avize" in cat_lower or "lamba" in cat_lower: return "Opviq"
    elif "furniture" in cat_lower or "mobilya" in cat_lower: return "Skye Decor"
    elif "bathroom" in cat_lower or "banyo" in cat_lower: return "Mijölnir"
    elif "bedroom" in cat_lower or "yatak odası" in cat_lower: return "L'Essentiel Linge de Maison"
    elif "decoration" in cat_lower or "dekorasyon" in cat_lower or "aksesuar" in cat_lower: return "Evila Originals"
    return ""

def validate_column_mappings(col_map, mappings):
    return [k for k in mappings if k not in col_map]

def process_wayfair_v19(data_file, template_file, ui_data, carton_file=None, progress_callback=None):
    data_file.seek(0)
    template_file.seek(0)
    
    df_data = pd.read_excel(data_file)
    if 'CODE' in df_data.columns:
        df_data = df_data.dropna(subset=['CODE'])
        df_data = df_data[df_data['CODE'].astype(str).str.strip() != '']
    else: 
        df_data = df_data.dropna(how='all')
        
    df_data = df_data.reset_index(drop=True)
    cat_col_name = next((col for col in df_data.columns if 'categor' in str(col).lower() or 'kategori' in str(col).lower()), None)
    
    carton_dict = {}
    if carton_file is not None:
        carton_file.seek(0)
        df_carton = pd.read_excel(carton_file)
        
        def find_col(df, keywords):
            for col in df.columns:
                if all(kw in str(col).lower() for kw in keywords): return col
            return None
            
        c_code_col = find_col(df_carton, ['code']) or find_col(df_carton, ['sku'])
        c_w_col = find_col(df_carton, ['weight'])
        c_x_col = find_col(df_carton, ['size', '- x'])
        c_y_col = find_col(df_carton, ['size', '- y'])
        c_z_col = find_col(df_carton, ['size', '- z'])
        
        if c_code_col:
            for _, r in df_carton.iterrows():
                c_sku = str(r[c_code_col]).strip()
                if c_sku and c_sku.lower() != 'nan':
                    if c_sku not in carton_dict: carton_dict[c_sku] = []
                    try:
                        w_val = float(r[c_w_col]) if c_w_col and pd.notna(r[c_w_col]) else 0
                        x_val = float(r[c_x_col]) if c_x_col and pd.notna(r[c_x_col]) else 0
                        y_val = float(r[c_y_col]) if c_y_col and pd.notna(r[c_y_col]) else 0
                        z_val = float(r[c_z_col]) if c_z_col and pd.notna(r[c_z_col]) else 0
                    except: 
                        w_val, x_val, y_val, z_val = 0, 0, 0, 0
                        
                    carton_dict[c_sku].append({'kg': w_val, 'x': x_val, 'y': y_val, 'z': z_val})

    wb = openpyxl.load_workbook(template_file)
    target_sheet = next((s for s in wb.sheetnames if not any(x in s for x in ["Additional", "WAYFAIR", "Instructions", "Valid Values", "Failed"])), wb.sheetnames[0])
    ws_main = wb[target_sheet]

    col_map = {}
    for c in range(1, ws_main.max_column + 1):
        r1_val = str(ws_main.cell(row=1, column=c).value).strip() if ws_main.cell(row=1, column=c).value else ""
        r4_val = str(ws_main.cell(row=4, column=c).value).strip() if ws_main.cell(row=4, column=c).value else ""
        col_let = ws_main.cell(row=1, column=c).column_letter
        
        if r1_val: col_map[r1_val] = col_let
            
        r4_lower = r4_val.lower()
        r1_lower = r1_val.lower()

        if ('color' in r4_lower or 'colour' in r4_lower or r1_lower.endswith('::color')):
            if 'leg' not in r4_lower and 'base' not in r4_lower and 'shade' not in r4_lower: col_map['featureDescription::color'] = col_let
        if 'overall height' in r4_lower or 'overallheight' in r1_lower: col_map['featureDescription::overallHeight'] = col_let
        elif 'overall width' in r4_lower or 'overallwidth' in r1_lower: col_map['featureDescription::overallWidth'] = col_let
        elif 'overall depth' in r4_lower or 'overalldepth' in r1_lower: col_map['featureDescription::overallDepth'] = col_let
        
        # Mapping for Overall Product Weight
        if 'overall product weight' in r4_lower or 'overallproductweight' in r1_lower: 
            col_map['featureDescription::overallProductWeight'] = col_let
            
        if 'set / single' in r4_lower: col_map['bedding::setSingle'] = col_let
        if 'bedding product type' in r4_lower: col_map['bedding::productType'] = col_let
        if 'bedding size' in r4_lower: col_map['bedding::size'] = col_let
        if 'bedding material' in r4_lower: col_map['bedding::material'] = col_let
        if 'pieces included' in r4_lower or 'total number of pieces included' in r4_lower: col_map['bedding::pieces'] = col_let

        for i in range(1, 6):
            if f'image file name or url {i}' in r4_lower: col_map[f'img_{i}'] = col_let

    feature_cols = [c.column_letter for c in ws_main[1] if str(c.value).strip() == 'featureDescription::genericFeatures']
    
    total_rows = len(df_data)
    processed, skipped, errors = 0, [], []
    missing_cols_reported = False
    written_rows, additional_images_data, additional_cartons_data = [], [], []
    processed_skus_for_additional = set()
    processed_skus_for_cartons = set()

    for index, row in df_data.iterrows():
        g_satir = 8 + index
        if progress_callback: progress_callback((index + 1) / total_rows)
            
        sku_key = str(row.get('CODE', '')).strip()
        try: pkg_count = int(float(row.get('NUMBER OF PACKAGES', 1)))
        except: pkg_count = 1

        raw_cartons = [{'kg': float(row.get('WEIGHT (Kg)', 0) or 0), 
                        'x': float(row.get('PACKAGING SIZE - X (cm)', 0) or 0),
                        'y': float(row.get('PACKAGING SIZE - Y (cm)', 0) or 0),
                        'z': float(row.get('PACKAGING SIZE - Z (cm)', 0) or 0)}]

        leave_carton_blank = False 

        if carton_file is not None and sku_key in carton_dict and len(carton_dict[sku_key]) > 0:
            raw_cartons = carton_dict[sku_key]
        elif pkg_count > 1:
            leave_carton_blank = True

        # ZORUNLU SIRALAMA: Hacim * Ağırlık prensibiyle en büyük koliyi 1. koli (Ana Koli) yapıyoruz
        raw_cartons.sort(key=lambda c: (c['x'] * c['y'] * c['z'], c['kg']), reverse=True)
        
        if len(raw_cartons) > 1 and not leave_carton_blank and sku_key not in processed_skus_for_cartons:
            for ext_c in raw_cartons[1:]: 
                additional_cartons_data.append({
                    'sku': sku_key, 'kg': ext_c['kg'], 'x': ext_c['x'], 'y': ext_c['y'], 'z': ext_c['z']
                })
            processed_skus_for_cartons.add(sku_key)

        kg = raw_cartons[0]['kg']
        x_cm = raw_cartons[0]['x']
        y_cm = raw_cartons[0]['y']
        z_cm = raw_cartons[0]['z']
        
        if len(raw_cartons) > 1:
            prod_weight_lbs = max(0, round((sum(c['kg'] for c in raw_cartons) * 2.20462) - 5, 2))
        else:
            prod_weight_lbs = max(0, round((kg - 0.1) * 2.20462, 2)) if kg > 0.1 else 0

        try:
            feat_text = row.get('FEATURES', '')
            raw_h, raw_w, raw_d = extract_overall_dims(feat_text)
            b_info = extract_bedding_info(feat_text, row.get('DESCRIPTION', ''), raw_h, raw_w)
            
            lbs = round(kg * 2.20462, 2)
            x_in = round(x_cm * 0.393701, 2)
            y_in = round(y_cm * 0.393701, 2)
            z_in = round(z_cm * 0.393701, 2)
            
            ean = row.get('EAN CODE', '')
            ean_str = "{:.0f}".format(float(ean)) if pd.notna(ean) and str(ean).strip() != '' else ""

            color_val = str(row.get('COLOR', ''))
            if color_val.lower() == 'nan': color_val = ''
            else: color_val = re.sub(r'\s*;\s*', '; ', color_val.replace('\n', ';').replace(',', ';').replace('/', ';')).strip('; ')

            cat_val = row.get(cat_col_name, '') if cat_col_name else ''
            auto_brand = get_brand_by_category(cat_val)

            mappings = {
                'core::supplierPartNumber': sku_key, 
                'core::manufacturerPartNumber': sku_key, 
                'core::universalProductCode': ean_str,
                'core::productName': b_info['new_name'] if b_info['new_name'] else row.get('DESCRIPTION'),
                'price::wholesalePrice': row.get('PRICE'), 
                'price::manufacturerSuggestedRetailPrice': row.get('RETAIL PRICE'),
                'featureDescription::overallHeight': convert_to_inch(raw_h) if ui_data['is_us'] else raw_h,
                'featureDescription::overallWidth': convert_to_inch(raw_w) if ui_data['is_us'] else raw_w,
                'featureDescription::overallDepth': convert_to_inch(raw_d) if ui_data['is_us'] else raw_d,
                'featureDescription::color': color_val, 
                'core::manufacturerId': auto_brand, 
                'shippingAndFulfillment::minimumOrderQuantity': 1, 
                'shippingAndFulfillment::forceQuantityMultiplier': 1, 
                'shippingAndFulfillment::displaySetQuantity': 1,
                'bedding::setSingle': b_info['set_single'], 
                'bedding::productType': b_info['prod_type'], 
                'bedding::size': b_info['bed_size'],
                'bedding::material': b_info['material'], 
                'bedding::pieces': b_info['pieces']
            }
            
            if leave_carton_blank:
                mappings['shippingAndFulfillment::weight'] = ""
                mappings['shippingAndFulfillment::height'] = ""
                mappings['shippingAndFulfillment::width'] = ""
                mappings['shippingAndFulfillment::depth'] = ""
                mappings['shippingAndFulfillment::productWeight'] = ""
                mappings['featureDescription::overallProductWeight'] = ""
            else:
                mappings['shippingAndFulfillment::weight'] = lbs
                mappings['shippingAndFulfillment::height'] = x_in
                mappings['shippingAndFulfillment::width'] = y_in
                mappings['shippingAndFulfillment::depth'] = z_in
                # Birebir Overall Product Weight ve normal weight'e yazılıyor
                mappings['shippingAndFulfillment::productWeight'] = prod_weight_lbs
                mappings['featureDescription::overallProductWeight'] = prod_weight_lbs

            urls = []
            for col in df_data.columns:
                col_str = str(col).lower()
                if 'image' in col_str or 'resim' in col_str or 'url' in col_str or 'link' in col_str:
                    if 'number' in col_str or 'sayı' in col_str or 'adet' in col_str: continue
                    val = str(row.get(col, '')).strip()
                    if val and val.lower() != 'nan' and (val.startswith('http') or val.startswith('www')) and val not in urls: 
                        urls.append(val)

            for i in range(min(5, len(urls))): mappings[f'img_{i+1}'] = urls[i]
                
            if len(urls) > 5 and sku_key not in processed_skus_for_additional:
                for ext_url in urls[5:]: additional_images_data.append((sku_key, ext_url))
                processed_skus_for_additional.add(sku_key)

            if ui_data['is_us']:
                mappings['shippingAndFulfillment::leadTime'] = 600
                mappings['shippingAndFulfillment::replacementLeadTime'] = 120
                
                if not leave_carton_blank and isinstance(x_in, (int, float)) and x_in > 0 and y_in > 0 and z_in > 0:
                    # KAPSAYICI LTL KONTROLÜ (Tüm kolileri tarıyoruz)
                    total_lbs = sum(c['kg'] for c in raw_cartons) * 2.20462
                    total_vol_in3 = sum((c['x'] * c['y'] * c['z']) for c in raw_cartons) * (0.393701 ** 3)
                    
                    is_ltl = False
                    for c in raw_cartons:
                        c_lbs = c['kg'] * 2.20462
                        c_l = c['x'] * 0.393701
                        c_w = c['y'] * 0.393701
                        c_h = c['z'] * 0.393701
                        dims = sorted([c_l, c_w, c_h], reverse=True)
                        length = dims[0]
                        girth = 2 * (dims[1] + dims[2])
                        if c_lbs >= 150 or (length + girth) >= 165 or length >= 108:
                            is_ltl = True
                            break
                            
                    fclass = calculate_freight_class_total(total_lbs, total_vol_in3)
                    
                    if is_ltl:
                        mappings['shippingAndFulfillment::shipType'] = "LTL"
                        mappings['shippingAndFulfillment::freightClass'] = fclass
                    else:
                        mappings['shippingAndFulfillment::shipType'] = "Small Parcel"

            if not missing_cols_reported:
                missing = validate_column_mappings(col_map, mappings)
                if missing: ui_data['missing_cols'] = missing
                missing_cols_reported = True

            for k, v in mappings.items():
                if k in col_map and pd.notna(v) and str(v).strip() != '': 
                    ws_main[f"{col_map[k]}{g_satir}"] = v

            for wid, val in ui_data['dyn_drops'].items():
                if wid in col_map and val:
                    if isinstance(val, list): final_str = "; ".join([str(pv) for pv in val if pv and str(pv) != 'None']) 
                    else: final_str = str(val)
                    if final_str: ws_main[f"{col_map[wid]}{g_satir}"] = final_str

            dim_writes = {
                'h': convert_to_inch(raw_h) if ui_data['is_us'] else raw_h, 
                'w': convert_to_inch(raw_w) if ui_data['is_us'] else raw_w, 
                'd': convert_to_inch(raw_d) if ui_data['is_us'] else raw_d
            }
            for dim_type, wids in ui_data['dim_mappings'].items():
                val = dim_writes[dim_type]
                if val is not None:
                    for wid in wids:
                        if wid in col_map: ws_main[f"{col_map[wid]}{g_satir}"] = val

            satirlar = [s.strip() for s in translate_features(feat_text, ui_data['is_us']).split('\n') if s.strip()]
            bedding_note = generate_bedding_note(feat_text, raw_h, raw_w, b_info['bed_size'], ui_data['is_us'])
            
            all_feats = satirlar.copy()
            if bedding_note: all_feats.append(bedding_note)
            n_feats = len(all_feats)
            
            # Öncelikle var olan Feature alanlarını temizleyelim
            for col_let in feature_cols: ws_main[f"{col_let}{g_satir}"] = ""

            if n_feats == 0:
                if len(feature_cols) > 0: ws_main[f"{feature_cols[0]}{g_satir}"] = "Made In Türkiye"
            elif n_feats <= 4:
                for i in range(n_feats):
                    if i < len(feature_cols): ws_main[f"{feature_cols[i]}{g_satir}"] = all_feats[i]
                if n_feats < len(feature_cols):
                    ws_main[f"{feature_cols[n_feats]}{g_satir}"] = "Made In Türkiye"
            else:
                for i in range(4):
                    if i < len(feature_cols): ws_main[f"{feature_cols[i]}{g_satir}"] = all_feats[i]
                if len(feature_cols) >= 5:
                    remaining_text = " | ".join(all_feats[4:])
                    ws_main[f"{feature_cols[4]}{g_satir}"] = f"{remaining_text} | Made In Türkiye"
                    
            processed += 1
            written_rows.append(g_satir)
            
        except Exception as e: 
            errors.append({'Satır': index + 2, 'Ürün Kodu': sku_key, 'Açıklama': str(row.get('DESCRIPTION', '') or '')[:60], 'Hata Detayı': str(e)})

    yellow_fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for c in range(1, ws_main.max_column + 1):
        if str(ws_main.cell(row=3, column=c).value).strip().lower() == "required":
            for r in written_rows:
                cell = ws_main.cell(row=r, column=c)
                if cell.value is None or str(cell.value).strip() == "": cell.fill = yellow_fill

    if additional_images_data:
        add_sheet = next((wb[s] for s in wb.sheetnames if 'additional' in s.lower() and 'image' in s.lower()), None)
        if add_sheet:
            sku_col_let, url_col_let, start_row = 'A', 'B', 5
            for c in range(1, add_sheet.max_column + 1):
                r1 = str(add_sheet.cell(row=1, column=c).value).lower()
                r4 = str(add_sheet.cell(row=4, column=c).value).lower()
                if 'supplier part number' in r4 or 'sku' in r4 or 'part number' in r1: sku_col_let = add_sheet.cell(row=1, column=c).column_letter
                if 'image file name or url' in r4 or 'url' in r4 or 'media::' in r1: url_col_let = add_sheet.cell(row=1, column=c).column_letter
            for r in range(4, add_sheet.max_row + 10):
                if not add_sheet[f"{sku_col_let}{r}"].value:
                    start_row = r
                    break
            for sku, url in additional_images_data:
                add_sheet[f"{sku_col_let}{start_row}"] = sku
                add_sheet[f"{url_col_let}{start_row}"] = url
                start_row += 1

    if additional_cartons_data:
        add_carton_sheet = next((wb[s] for s in wb.sheetnames if 'additional' in s.lower() and ('carton' in s.lower() or 'package' in s.lower())), None)
        if add_carton_sheet:
            col_map_c = {}
            for c in range(1, add_carton_sheet.max_column + 1):
                r1, r4, let = str(add_carton_sheet.cell(row=1, column=c).value).lower(), str(add_carton_sheet.cell(row=4, column=c).value).lower(), add_carton_sheet.cell(row=1, column=c).column_letter
                if 'supplier part number' in r4 or 'sku' in r4 or 'part number' in r1: col_map_c['sku'] = let
                elif 'weight' in r4 or 'weight' in r1: col_map_c['weight'] = let
                elif 'height' in r4 or 'height' in r1: col_map_c['height'] = let
                elif 'width' in r4 or 'width' in r1: col_map_c['width'] = let
                elif 'depth' in r4 or 'depth' in r1: col_map_c['depth'] = let
                    
            if 'sku' in col_map_c:
                start_row = 5
                for r in range(4, add_carton_sheet.max_row + 10):
                    if not add_carton_sheet[f"{col_map_c['sku']}{r}"].value:
                        start_row = r; break
                        
                for c_data in additional_cartons_data:
                    c_w_final = round(c_data['kg'] * 2.20462, 2) if ui_data['is_us'] else c_data['kg']
                    c_h_final = round(c_data['x'] * 0.393701, 2) if ui_data['is_us'] else c_data['x']
                    c_w_final_2 = round(c_data['y'] * 0.393701, 2) if ui_data['is_us'] else c_data['y']
                    c_d_final = round(c_data['z'] * 0.393701, 2) if ui_data['is_us'] else c_data['z']
                    
                    add_carton_sheet[f"{col_map_c['sku']}{start_row}"] = c_data['sku']
                    if 'weight' in col_map_c: add_carton_sheet[f"{col_map_c['weight']}{start_row}"] = c_w_final
                    if 'height' in col_map_c: add_carton_sheet[f"{col_map_c['height']}{start_row}"] = c_h_final
                    if 'width' in col_map_c: add_carton_sheet[f"{col_map_c['width']}{start_row}"] = c_w_final_2
                    if 'depth' in col_map_c: add_carton_sheet[f"{col_map_c['depth']}{start_row}"] = c_d_final
                    start_row += 1

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue(), processed, skipped, errors

def process_data_excel_only(data_file, is_us):
    data_file.seek(0)
    wb = openpyxl.load_workbook(data_file)
    ws = wb.active

    if hasattr(ws, '_images'): ws._images = []

    def get_headers():
        h = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=1, column=c).value
            if v: h[str(v).strip()] = c
        return h

    headers = get_headers()
    code_col = headers.get('CODE')
    if code_col:
        for row in range(ws.max_row, 1, -1):
            val = ws.cell(row=row, column=code_col).value
            if val is None or str(val).strip() == '': ws.delete_rows(row, 1)
                
    headers = get_headers()
    ean_col = headers.get('EAN CODE')
    if ean_col:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=ean_col)
            if cell.value is not None:
                try: cell.value = int(float(str(cell.value).strip())); cell.number_format = '0'
                except: cell.value = str(cell.value).strip(); cell.number_format = '@'

    color_col = headers.get('COLOR')
    if color_col:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=color_col)
            if cell.value:
                c_val = str(cell.value)
                if c_val.lower() != 'nan':
                    c_val = c_val.replace('\n', ';').replace(',', ';').replace('/', ';')
                    cell.value = re.sub(r'\s*;\s*', '; ', c_val).strip('; ')

    ef_col = headers.get('EXTRA FEATURES')
    if ef_col:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=ef_col)
            if cell.value: cell.value = translate_features(str(cell.value), is_us)

    feat_col = headers.get('FEATURES')
    if feat_col:
        ws.insert_cols(feat_col + 1, 5)
        ref_header = ws.cell(row=1, column=feat_col)
        ref_col_letter = get_column_letter(feat_col)
        ref_width = ws.column_dimensions[ref_col_letter].width
        
        for i, f_name in enumerate(['Feature 1', 'Feature 2', 'Feature 3', 'Feature 4', 'Feature 5']):
            target_col_idx = feat_col + 1 + i
            target_col_letter = get_column_letter(target_col_idx)
            nc = ws.cell(row=1, column=target_col_idx)
            nc.value = f_name
            if ref_width: ws.column_dimensions[target_col_letter].width = ref_width
            if ref_header.has_style:
                nc.font = copy.copy(ref_header.font); nc.border = copy.copy(ref_header.border)
                nc.fill = copy.copy(ref_header.fill); nc.alignment = copy.copy(ref_header.alignment)

        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=feat_col)
            features_to_write = ["", "", "", "", ""]
            
            if cell.value:
                translated = translate_features(str(cell.value), is_us)
                lines = [s.strip() for s in translated.split('\n') if s.strip()]
            else:
                lines = []
                
            n = len(lines)
            if n == 0:
                features_to_write[0] = "Made In Türkiye"
            elif n <= 4:
                for idx in range(n): features_to_write[idx] = lines[idx]
                features_to_write[n] = "Made In Türkiye"
            else:
                for idx in range(4): features_to_write[idx] = lines[idx]
                remaining_text = " | ".join(lines[4:])
                features_to_write[4] = f"{remaining_text} | Made In Türkiye"

            ws.cell(row=row, column=feat_col + 1).value = features_to_write[0]
            ws.cell(row=row, column=feat_col + 2).value = features_to_write[1]
            ws.cell(row=row, column=feat_col + 3).value = features_to_write[2]
            ws.cell(row=row, column=feat_col + 4).value = features_to_write[3]
            ws.cell(row=row, column=feat_col + 5).value = features_to_write[4]

    if is_us:
        headers = get_headers()
        w_col, x_col, y_col, z_col = headers.get('WEIGHT (Kg)'), headers.get('PACKAGING SIZE - X (cm)'), headers.get('PACKAGING SIZE - Y (cm)'), headers.get('PACKAGING SIZE - Z (cm)')
        metric_cols = [c for c in [w_col, x_col, y_col, z_col] if c is not None]
        if metric_cols:
            insert_idx = max(metric_cols) + 1
            ws.insert_cols(insert_idx, 4)
            ws.cell(row=1, column=insert_idx).value = 'WEIGHT (Lbs)'
            ws.cell(row=1, column=insert_idx + 1).value = 'PACKAGING SIZE - X (in)'
            ws.cell(row=1, column=insert_idx + 2).value = 'PACKAGING SIZE - Y (in)'
            ws.cell(row=1, column=insert_idx + 3).value = 'PACKAGING SIZE - Z (in)'
            
            ref_h = ws.cell(row=1, column=metric_cols[0])
            for i in range(4):
                nh = ws.cell(row=1, column=insert_idx + i)
                if ref_h.has_style:
                    nh.font = copy.copy(ref_h.font); nh.border = copy.copy(ref_h.border)
                    nh.fill = copy.copy(ref_h.fill); nh.alignment = copy.copy(ref_h.alignment)
                    
            for row in range(2, ws.max_row + 1):
                if w_col:
                    val = ws.cell(row=row, column=w_col).value
                    try: ws.cell(row=row, column=insert_idx).value = round(float(val) * 2.20462, 2)
                    except: pass
                if x_col:
                    val = ws.cell(row=row, column=x_col).value
                    try: ws.cell(row=row, column=insert_idx + 1).value = round(float(val) * 0.393701, 2)
                    except: pass
                if y_col:
                    val = ws.cell(row=row, column=y_col).value
                    try: ws.cell(row=row, column=insert_idx + 2).value = round(float(val) * 0.393701, 2)
                    except: pass
                if z_col:
                    val = ws.cell(row=row, column=z_col).value
                    try: ws.cell(row=row, column=insert_idx + 3).value = round(float(val) * 0.393701, 2)
                    except: pass

    headers = get_headers()
    img_cols = []
    for col_name, col_idx in headers.items():
        name_lower = col_name.lower()
        if 'image' in name_lower or 'resim' in name_lower or 'url' in name_lower or 'link' in name_lower: img_cols.append((col_idx, col_name))
            
    if img_cols:
        img_cols.sort(key=lambda x: x[0], reverse=True)
        extracted_cols = []
        for col_idx, col_name in img_cols:
            col_data = []
            col_letter = get_column_letter(col_idx)
            col_width = ws.column_dimensions[col_letter].width
            for row in range(1, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                cell_data = {
                    'value': cell.value, 'font': copy.copy(cell.font) if cell.has_style and cell.font else None,
                    'border': copy.copy(cell.border) if cell.has_style and cell.border else None,
                    'fill': copy.copy(cell.fill) if cell.has_style and cell.fill else None,
                    'alignment': copy.copy(cell.alignment) if cell.has_style and cell.alignment else None, 'number_format': cell.number_format
                }
                col_data.append(cell_data)
            extracted_cols.append({'name': col_name, 'width': col_width, 'data': col_data})
            ws.delete_cols(col_idx, 1)
            
        extracted_cols.reverse()
        for col_dict in extracted_cols:
            new_col_idx = ws.max_column + 1
            new_col_letter = get_column_letter(new_col_idx)
            if col_dict['width']: ws.column_dimensions[new_col_letter].width = col_dict['width']
            for row_idx, c_data in enumerate(col_dict['data'], start=1):
                new_cell = ws.cell(row=row_idx, column=new_col_idx)
                new_cell.value = c_data['value']
                if c_data['font']: new_cell.font = c_data['font']
                if c_data['border']: new_cell.border = c_data['border']
                if c_data['fill']: new_cell.fill = c_data['fill']
                if c_data['alignment']: new_cell.alignment = c_data['alignment']
                if c_data['number_format']: new_cell.number_format = c_data['number_format']

    medium_border = Border(left=Side(style='medium', color='000000'), right=Side(style='medium', color='000000'), top=Side(style='medium', color='000000'), bottom=Side(style='medium', color='000000'))
    column_colors = {
        'Feature 1': "DDEBF7", 'Feature 2': "E2EFDA", 'Feature 3': "FFF2CC", 'Feature 4': "FCE4D6", 'Feature 5': "E8D8FC",
        'WEIGHT (Lbs)': "F8CECC", 'PACKAGING SIZE - X (in)': "D1F2EB", 'PACKAGING SIZE - Y (in)': "E8F8F5", 'PACKAGING SIZE - Z (in)': "E6F2F7"
    }
    
    headers = get_headers()
    for col_name, col_idx in headers.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 12
        col_color_hex = column_colors.get(col_name)
        col_fill = PatternFill(start_color=col_color_hex, end_color=col_color_hex, fill_type="solid") if col_color_hex else None
            
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.border = medium_border
            if row == 1: cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='justify', vertical='center', wrap_text=True)
                is_bold = cell.font.bold if (cell.font and cell.font.bold is not None) else False
                is_italic = cell.font.italic if (cell.font and cell.font.italic is not None) else False
                cell.font = Font(name='Tahoma', size=8, bold=is_bold, italic=is_italic)
                if col_fill: cell.fill = col_fill

    for row in range(1, ws.max_row + 1): ws.row_dimensions[row].height = 18

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

st.set_page_config(page_title="Wayfair & Data Akıllı Ürün Robotu V19", layout="wide")
st.title("🛡️ Wayfair & Data Akıllı Ürün Robotu V19")

with st.sidebar:
    st.header("⚙️ Genel Ayarlar")
    region_selection = st.radio("🌎 Bölge Seçimi", ["US (İnç / Lbs)", "EU (cm / Kg)"], horizontal=False)
    is_us = region_selection.startswith("US") 
    st.divider()
    st.subheader("🔗 Hızlı Bağlantılar")
    st.markdown("[🛠️ Asir Tools](https://excelwebpy-asirtools.streamlit.app/)")
    st.divider()
    
tab_wayfair, tab_data = st.tabs(["🎯 Wayfair Şablonu Hazırla", "🛠️ Sadece Data Excel'i Çevir"])

with tab_wayfair:
    st.info("💡 Bu alan, Data Excel'inizdeki ürünleri Wayfair şablonuna otomatik eşleştirip yazar.")
    
    u1, u2, u3 = st.columns(3)
    with u1: d_file = st.file_uploader("1. Data Excel", type="xlsx", key="wayfair_data")
    with u2: t_file = st.file_uploader("2. Template Excel", type="xlsx", key="wayfair_template")
    with u3: c_file = st.file_uploader("3. Paket Excel (Opsiyon)", type="xlsx", key="wayfair_carton")

    AUTO_MAPPED_COLS = {
        'core::supplierPartNumber', 'core::manufacturerPartNumber', 'core::universalProductCode',
        'core::productName', 'featureDescription::romanceCopy', 'featureDescription::overallHeight',
        'featureDescription::overallWidth', 'featureDescription::overallDepth', 'featureDescription::color',
        'featureDescription::genericFeatures', 'shippingAndFulfillment::weight', 'shippingAndFulfillment::height',
        'shippingAndFulfillment::width', 'shippingAndFulfillment::depth', 'price::wholesalePrice',
        'price::manufacturerSuggestedRetailPrice', 'shippingAndFulfillment::minimumOrderQuantity',
        'shippingAndFulfillment::forceQuantityMultiplier', 'shippingAndFulfillment::displaySetQuantity',
        'shippingAndFulfillment::productWeight', 'featureDescription::overallProductWeight', 'shippingAndFulfillment::leadTime',
        'shippingAndFulfillment::replacementLeadTime', 'shippingAndFulfillment::shipType',
        'shippingAndFulfillment::freightClass', 'core::collectionName', 'core::manufacturerId',
        'featureDescription::marketingCopy', 'bedding::setSingle', 'bedding::productType', 
        'bedding::size', 'bedding::material', 'bedding::pieces'
    }

    def is_auto_mapped_by_fname(fname):
        f_low = fname.lower().strip()
        exact_matches = {
            'overall height', 'overall width', 'overall depth', 'overallheight', 
            'overallwidth', 'overalldepth', 'overall product weight', 'overallproductweight', 'color', 'colour', 'marketing copy', 
            'marketingcopy', 'set / single', 'bedding product type', 'bedding size', 
            'bedding material', 'pieces included', 'total number of pieces included'
        }
        return f_low in exact_matches

    if d_file and t_file:
        t_bytes = t_file.getvalue()
        try: df_v = pd.read_excel(io.BytesIO(t_bytes), sheet_name='Valid Values')
        except: df_v = None

        wb_t = openpyxl.load_workbook(io.BytesIO(t_bytes))
        target_name = next((s for s in wb_t.sheetnames if not any(x in s for x in ["Additional", "WAYFAIR", "Instructions", "Valid Values", "Failed"])), wb_t.sheetnames[0])
        ws_t = wb_t[target_name]

        eligible_cols = []
        for c in range(1, ws_t.max_column + 1):
            wid = str(ws_t.cell(1, c).value).strip()
            status = str(ws_t.cell(3, c).value).strip()
            fname = str(ws_t.cell(4, c).value).strip()
            if status.lower() == "required" and wid not in AUTO_MAPPED_COLS and not wid.startswith('media::') and not is_auto_mapped_by_fname(fname):
                eligible_cols.append((wid, fname))

        st.markdown("---")
        st.subheader("📐 Özel Ölçü Sütun Eşleştirmeleri")
        options_dict = {f"{fname} ({wid})": wid for wid, fname in eligible_cols}
        options_list = list(options_dict.keys())
        
        col_h, col_w, col_d = st.columns(3)
        with col_h: h_sel = st.multiselect("Height (Yükseklik) Yazılacaklar", options=options_list)
        with col_w: w_sel = st.multiselect("Genişlik (Width) Yazılacaklar", options=options_list)
        with col_d: d_sel = st.multiselect("Depth (Derinlik) Yazılacaklar", options=options_list)

        selected_dim_wids = [options_dict[x] for x in h_sel + w_sel + d_sel]
        dim_mappings = {'h': [options_dict[x] for x in h_sel], 'w': [options_dict[x] for x in w_sel], 'd': [options_dict[x] for x in d_sel]}

        st.markdown("---")
        st.subheader(f"📋 {target_name} — Doldurulması Gereken Diğer Özellikler")

        dyn_selections = {}
        cols_ui = st.columns(3)
        idx = 0

        for wid, fname in eligible_cols:
            if wid in selected_dim_wids: continue
                
            with cols_ui[idx % 3]:
                if df_v is not None and fname in df_v.columns:
                    opts = list(dict.fromkeys([str(o).strip() for o in df_v[fname].dropna().unique() if str(o).strip() and str(o).strip() != 'None']))
                    # GÜNCELLEME: "Select all" seçeneklerini ortadan kaldırıyoruz
                    opts = [o for o in opts if "select all" not in o.lower()]
                else: 
                    opts = ["Yes", "No", "Does Not Apply"]
                    
                # GÜNCELLEME: Custom değer için seçenek ekleme
                opts.append("➕ Custom Value")

                f_low = fname.lower()
                if wid not in st.session_state['user_prefs']:
                    def_val = []
                    if 'warning required' in f_low: def_val = ['No']
                    elif 'country of manufacturer' in f_low or 'country of origin' in f_low: 
                        # GÜNCELLEME: Default Country -> Turkey
                        def_val = ['Turkey'] if 'Turkey' in opts else (['Türkiye'] if 'Türkiye' in opts else [])
                    elif 'uniform packaging and labeling regulations' in f_low: def_val = ['Yes']
                    elif 'reason for restriction' in f_low: def_val = ['Does Not Apply']
                    elif 'general certificate of conformity' in f_low: def_val = ['Yes']
                    elif 'canada product restriction' in f_low: def_val = ['No']
                    elif 'soffa compliant' in f_low: def_val = ['Does Not Apply']
                    elif 'canfer compliant' in f_low: def_val = ['Does Not Apply']
                    elif 'carb phase' in f_low: def_val = ['Does Not Apply']
                    elif 'composite wood product (cwp)' in f_low: def_val = ['Does Not Apply']
                    elif 'tsca title vi compliant' in f_low: def_val = ['Does Not Apply']
                    elif 'supplier intended and approved use' in f_low:
                        def_val = [x for x in ['Non Residential Use', 'Residential Use'] if x in opts]
                        if not def_val: def_val = ['Non Residential Use', 'Residential Use']  
                    elif 'commercial warranty' in f_low: def_val = ['Yes'] 
                    elif 'contains flame retardant' in f_low: def_val = ['No']
                    elif 'wayfair compliance verified' in f_low: def_val = ['No']
                    elif 'battery or batteries included' in f_low: def_val = ['No']
                    elif 'additional intended use for child' in f_low: def_val = ['No']
                    st.session_state['user_prefs'][wid] = def_val

                saved = st.session_state['user_prefs'].get(wid, [])
                sel = st.multiselect(fname, options=opts, default=[x for x in saved if x in opts], key=f"sel_{wid}")
                
                # GÜNCELLEME: Eğer Custom Value seçildiyse text box göster
                final_sel = list(sel)
                if "➕ Custom Value" in final_sel:
                    custom_val = st.text_input(f"✍️ {fname} için Custom Değer Girin:", key=f"custom_{wid}")
                    final_sel.remove("➕ Custom Value")
                    if custom_val:
                        final_sel.append(custom_val)
                
                dyn_selections[wid] = final_sel
                st.session_state['user_prefs'][wid] = sel
                
            idx += 1

        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🚀 Wayfair Dosyasını Hazırla", type="primary", width='stretch'):
            ui_data = {'is_us': is_us, 'dyn_drops': dyn_selections, 'dim_mappings': dim_mappings, 'missing_cols': []}
            progress_bar = st.progress(0, text="Hazırlanıyor...")
            def update_progress(val): progress_bar.progress(min(val, 1.0), text=f"İşleniyor... %{int(val * 100)}")

            with st.spinner("Excel dosyası işleniyor..."):
                d_io, t_io = io.BytesIO(d_file.getvalue()), io.BytesIO(t_file.getvalue())
                c_io = io.BytesIO(c_file.getvalue()) if c_file else None
                res, processed, skipped, errors = process_wayfair_v19(d_io, t_io, ui_data, carton_file=c_io, progress_callback=update_progress)

            progress_bar.progress(1.0, text="✅ Tamamlandı!")
            
            st.markdown("---")
            m1, m2, m3, m4 = st.columns(4)
            m1.metric("✅ İşlenen", processed); m2.metric("⏭️ Atlanan", len(skipped))
            m3.metric("❌ Hatalı", len(errors)); m4.metric("📦 Toplam", processed + len(skipped) + len(errors))

            if ui_data.get('missing_cols'):
                with st.expander(f"⚠️ {len(ui_data['missing_cols'])} Sütun Template'de Bulunamadı", expanded=True):
                    st.warning("Bu sütunlar mapping'de tanımlı ama template'de yok — ilgili veriler yazılamadı:")
                    st.code("\n".join(ui_data['missing_cols']))
            
            if skipped:
                with st.expander(f"⏭️ Atlanan Satırlar — {len(skipped)} ürün", expanded=False): st.dataframe(pd.DataFrame(skipped), width='stretch')
            if errors:
                with st.expander(f"❌ Hatalı Satırlar — {len(errors)} ürün", expanded=True):
                    st.error("Hata Detayları:")
                    st.dataframe(pd.DataFrame(errors), width='stretch')
            
            if processed > 0:
                st.success(f"✅ {processed} ürün başarıyla işlendi.")
                st.download_button(label="📥 Hazır Excel'i İndir", data=res, file_name="Wayfair_Upload-Template.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

with tab_data:
    st.subheader("🛠️ Data Excel Dönüştürücü (Şablonsuz)")
    st.info("💡 Bu araç sadece veri excelinizi okur, **özellikleri (Features) 5 sütuna böler**, ölçüleri (cm/kg/ml) Amerikan standartlarına (inç/lbs/fl oz) çevirir ve koli ebatları için yeni sütunlar açıp size temiz bir Excel sunar.")
    
    data_only_file = st.file_uploader("İşlenecek Data Excel'ini Yükleyin", type="xlsx", key="data_only_upload")
    if data_only_file:
        if st.button("🚀 Data Excel'i Dönüştür ve İndir", type="primary", width='stretch'):
            with st.spinner("Data Excel'iniz çevriliyor ve bölünüyor..."):
                d_io = io.BytesIO(data_only_file.getvalue())
                result_excel = process_data_excel_only(d_io, is_us)
                
                st.success("✅ Dönüştürme Başarılı! Aşağıdaki butona tıklayarak yeni excelinizi indirebilirsiniz.")
                st.download_button(label="📥 Dönüştürülmüş Data Excel'i İndir", data=result_excel, file_name="Converted_Data_Excel.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

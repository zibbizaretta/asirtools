import pandas as pd
import math
import re
import io
import os
from datetime import datetime
import streamlit as st
import pypdf
from openpyxl.styles import Alignment

# --- CONSTANTS ---
KG_TO_LBS = 2.20462
CM_TO_INCH = 0.393701
MADE_IN_TURKEY = "Made In Türkiye"

# --- CONSTANTS (TITLE GENERATOR TOOL) ---
COLORS = [
    "Black", "White", "Gold", "Silver", "Anthracite", "Walnut", "Oak", "Beige", 
    "Ecru", "Turquoise", "Pink", "Red", "Blue", "Green", "Yellow", "Orange", 
    "Purple", "Grey", "Gray", "Brown", "Copper", "Bronze", "Chrome", "Mustard", 
    "Fuchsia", "Mink", "Salmon", "Claret Red", "Navy Blue", "Mint"
]

MATERIALS = [
    "MDF", "Metal", "Steel", "Tempered Glass", "Glass", "Velvet", "Cotton", 
    "Spruce Wood", "Pine Wood", "Beech Wood", "Solid Wood", "Wood", "Plastic", 
    "Ceramic", "Polyester", "Acrylic", "Jute", "Coir", "PVC", "Bamboo", "Leather", 
    "Linen", "Porcelain", "Iron"
]

# --- HELPER FUNCTIONS (GENERAL & WF TEMPLATE TOOL) ---
def extract_dimensions_from_string(text_to_search):
    def find_dimension_value(pattern, text):
        match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
        if match:
            try:
                value_str = match.group(1).replace(',', '.')
                return float(value_str)
            except: return None
        return None
    w = find_dimension_value(r'(?:Width|Genişlik):\s*(\d+(?:[.,]\d+)?)', text_to_search)
    h = find_dimension_value(r'(?:Height|Yükseklik):\s*(\d+(?:[.,]\d+)?)', text_to_search)
    d = find_dimension_value(r'(?:Depth|Derinlik):\s*(\d+(?:[.,]\d+)?)', text_to_search)
    l = find_dimension_value(r'(?:Length|Uzunluk):\s*(\d+(?:[.,]\d+)?)', text_to_search)
    diam = find_dimension_value(r'(?:Diameter|Çap):\s*(\d+(?:[.,]\d+)?)', text_to_search)
    y_val = d if d is not None else l
    if w is not None and h is not None and y_val is not None: return (w, y_val, h)
    if diam is not None and h is not None: return (diam, diam, h)
    xyz_pattern = r'(\d+(?:[.,]\d+)?)\s*[xX]\s*(\d+(?:[.,]\d+)?)(?:\s*[xX]\s*(\d+(?:[.,]\d+)?))?'
    match = re.search(xyz_pattern, text_to_search)
    if match:
        try:
            x = float(match.group(1).replace(',', '.'))
            y = float(match.group(2).replace(',', '.'))
            z = float(match.group(3).replace(',', '.')) if match.group(3) else None
            return (x, y, z)
        except: return None
    return None

def clean_feature_list(features_str):
    if pd.isna(features_str) or features_str == "": return []
    features = re.split(r'\s*(?:\\n|\n)\s*', str(features_str).strip())
    return [f.strip() for f in features if f and f.strip()]

def convert_size_value(val, unit_choice):
    if val is None or val == '' or (isinstance(val, float) and math.isnan(val)): return ''
    try:
        num_val = float(str(val).replace(',', '.'))
        return round(num_val * CM_TO_INCH, 2) if unit_choice == "inch" else round(num_val, 2)
    except: return val

def convert_weight_value(val_kg, weight_unit_choice):
    if val_kg is None or val_kg == '' or (isinstance(val_kg, float) and math.isnan(val_kg)): return ''
    try:
        num_val = float(str(val_kg).replace(',', '.'))
        if weight_unit_choice == "LBS":
            c_lbs = round(num_val * KG_TO_LBS, 2)
            return c_lbs, max(0.0, round(c_lbs - 0.01, 2))
        return round(num_val, 2), round(num_val, 2)
    except: return val_kg, val_kg


# --- HELPER FUNCTIONS (PO TRACKING TOOL) ---
def process_pdfs_robust(pdf_files):
    all_data = {}
    po_pattern = re.compile(r"((?:CS|CA)\d{9,})")
    trk_pattern = re.compile(r"(?<![a-zA-Z0-9])(\d{12,20})(?![a-zA-Z0-9])")
    
    for pdf_file in pdf_files:
        try:
            reader = pypdf.PdfReader(pdf_file)
            for page in reader.pages:
                text = page.extract_text()
                if not text: continue
                
                chunks = po_pattern.split(text)
                for j in range(1, len(chunks), 2):
                    po_number = chunks[j].strip()
                    content_after_po = chunks[j+1]
                    
                    found_trks = trk_pattern.findall(content_after_po)
                    if found_trks:
                        if po_number not in all_data: all_data[po_number] = set()
                        for t in found_trks:
                            if not t.startswith("26"):
                                all_data[po_number].add(t)
        except Exception as e:
            st.error(f"Error: {pdf_file.name} - {e}")
    
    final_rows = []
    for po in sorted(all_data.keys()):
        trks = sorted(list(all_data[po]))
        if trks:
            final_rows.append({"PO": po, "TRK": ", ".join(trks)})
    return pd.DataFrame(final_rows)


# --- HELPER FUNCTIONS (TITLE GENERATOR TOOL) ---
def clean_title_text(text):
    """Gereksiz kısımları metinden atar."""
    if not isinstance(text, str): return ""
    
    ignore_keywords = [
        r"number of packages", r"paket sayısı", r"packaging size", 
        r"weight", r"ean code", r"price", r"retail price"
    ]
    
    for keyword in ignore_keywords:
        match = re.search(keyword, text, flags=re.IGNORECASE)
        if match:
            text = text[:match.start()]
            
    return text.strip()

def extract_color(text):
    found_colors = []
    for color in COLORS:
        if re.search(rf"\b{color}\b", text, re.IGNORECASE):
            if color not in found_colors:
                found_colors.append(color)
    return " & ".join(found_colors[:2])

def extract_material(text):
    found_materials = []
    percent_match = re.search(r"100%\s*([a-zA-Z\s]+?)(?:,|\||Fabric|Construction|Frame)", text, re.IGNORECASE)
    if percent_match:
        mat = percent_match.group(1).strip()
        if len(mat.split()) <= 3:
            return mat.title()

    for mat in MATERIALS:
        if re.search(rf"\b{mat}\b", text, re.IGNORECASE):
            if mat not in found_materials:
                found_materials.append(mat)
    return " & ".join(found_materials[:2])

def extract_and_convert_title_dimensions(text, target_unit):
    text = text.replace(',', '.')
    
    pattern_abc = r'(\d+(?:\.\d+)?)\s*(?:x|X|\*)\s*(\d+(?:\.\d+)?)(?:\s*(?:x|X|\*)\s*(\d+(?:\.\d+)?))?\s*(cm|in|inch|inches|mm)?'
    match = re.search(pattern_abc, text, re.IGNORECASE)
    
    pattern_whd = r'(?:W|Width|Genişlik)\s*:?\s*(\d+(?:\.\d+)?).*?(?:H|Height|Yükseklik)\s*:?\s*(\d+(?:\.\d+)?)'
    match_whd = re.search(pattern_whd, text, re.IGNORECASE)

    val1, val2, val3, current_unit = None, None, None, "cm"

    if match:
        val1, val2 = float(match.group(1)), float(match.group(2))
        val3 = float(match.group(3)) if match.group(3) else None
        if match.group(4):
            current_unit = match.group(4).lower()
    elif match_whd:
        val1, val2 = float(match_whd.group(1)), float(match_whd.group(2))
        if "inch" in text.lower() or "in" in text.lower().split():
            current_unit = "inch"

    if val1 is None: return ""

    def convert(v, c_unit, t_unit):
        if "in" in c_unit and t_unit == "cm": return v * 2.54
        if "cm" in c_unit and t_unit == "inch": return v / 2.54
        if "mm" in c_unit and t_unit == "inch": return v / 25.4
        if "mm" in c_unit and t_unit == "cm": return v / 10
        return v

    nv1 = round(convert(val1, current_unit, target_unit), 1)
    nv2 = round(convert(val2, current_unit, target_unit), 1)
    nv3 = round(convert(val3, current_unit, target_unit), 1) if val3 else None

    unit_str = "inches" if target_unit == "inch" else "cm"
    
    if nv3: return f"{nv1} x {nv2} x {nv3} {unit_str}"
    else: return f"{nv1} x {nv2} {unit_str}"

def generate_title(row, name_col, features_col, target_unit, marketing_col=None):
    raw_name = str(row[name_col]) if pd.notna(row[name_col]) else ""
    raw_features = str(row[features_col]) if features_col and pd.notna(row[features_col]) else ""
    
    full_text = clean_title_text(raw_name + " | " + raw_features)
    
    base_name = ""
    if marketing_col and marketing_col in row and pd.notna(row[marketing_col]):
        base_name = str(row[marketing_col]).strip()
    else:
        base_name = re.split(r',|\|', raw_name)[0].strip()

    color = extract_color(full_text)
    material = extract_material(full_text)
    dimensions = extract_and_convert_title_dimensions(full_text, target_unit)

    title_parts = [base_name]
    if color: title_parts.append(color)
    if material: title_parts.append(material)
    if dimensions: title_parts.append(dimensions)

    return " - ".join(title_parts)


# --- APP INTERFACE ---
st.set_page_config(page_title="Asir Tools", layout="wide")

with st.sidebar:
    st.title("Asir Tools")
    # YENİ TOOL BURAYA EKLENDİ
    page = st.radio("Select Tool:", ["WF Template Tool", "PO Tracking Tool", "Title Generator Tool"])
    st.divider()
    if st.button("Home / Reset", use_container_width=True):
        st.write('<meta http-equiv="refresh" content="0;url=https://excelwebpy-asirtools.streamlit.app/">', unsafe_allow_html=True)
        st.stop()


# --- PAGE 1: WF TEMPLATE TOOL ---
if page == "WF Template Tool":
    st.header("WF Template Tool")
    with st.sidebar:
        st.subheader("Settings")
        size_unit = st.radio("Size Unit:", ("cm", "inch"), index=1)
        weight_unit = st.radio("Weight Unit:", ("KG", "LBS"), index=1)
        add_made_in_tr = st.checkbox("Add 'Made in Türkiye'", value=True)
        feature_count = st.slider("Feature Column Count:", 1, 10, 5)

    uploaded_file = st.file_uploader("Upload Excel file", type=["xlsx", "xls"])
    if uploaded_file:
        try:
            file_base, file_ext = os.path.splitext(uploaded_file.name)
            output_filename = f"{file_base}_processed{file_ext}"
            df = pd.read_excel(uploaded_file, dtype=str).fillna('')
            processed_data = []
            
            f_headers = [f'Feature {i+1}' for i in range(feature_count)]
            u_s, u_w = f"({size_unit})", f"({weight_unit})"
            output_headers = ['CODE', 'EAN CODE', 'COLOR', 'DESCRIPTION'] + f_headers + \
                             ['IMAGE', 'PRICE', ' ', 'RETAIL PRICE', 'NUMBER OF PACKAGES', f'WEIGHT {u_w}',
                              f'PRODUCT SIZE - X {u_s}', f'PRODUCT SIZE - Y {u_s}', f'PRODUCT SIZE - Z {u_s}',
                              f'CARTON WEIGHT {u_w}', f'PACKAGING SIZE - X {u_s}', f'PACKAGING SIZE - Y {u_s}', f'PACKAGING SIZE - Z {u_s}']

            for index, row in df.iterrows():
                code = str(row.get('CODE', '')).strip()
                if not code or code.lower() == 'nan': continue
                feat_list = clean_feature_list(row.get('FEATURES', ''))
                extra = str(row.get('EXTRA FEATURES', ''))
                if "number of packages" not in extra.lower(): feat_list.extend(clean_feature_list(extra))
                if add_made_in_tr and not any(MADE_IN_TURKEY in str(f) for f in feat_list): feat_list.append(MADE_IN_TURKEY)
                
                f_cols = [""] * feature_count
                if feature_count > 1:
                    for i in range(min(len(feat_list), feature_count - 1)): f_cols[i] = feat_list[i]
                    if len(feat_list) >= feature_count: f_cols[feature_count-1] = "\n".join(feat_list[feature_count-1:])
                elif feat_list: f_cols[0] = "\n".join(feat_list)

                dims = extract_dimensions_from_string(str(row.get('FEATURES', '')) + "\n" + extra)
                px, py, pz = dims if dims else ('', '', '')
                c_w, p_w = convert_weight_value(row.get('WEIGHT (Kg)', ''), weight_unit)

                processed_data.append([code, row.get('EAN CODE', ''), str(row.get('COLOR', '')).replace('\n', ';'), row.get('DESCRIPTION', '')] + f_cols + \
                                      [row.get('IMAGE', ''), row.get('PRICE', ''), '', row.get('RETAIL PRICE', ''), row.get('NUMBER OF PACKAGES', ''),
                                       p_w, convert_size_value(px, size_unit), convert_size_value(py, size_unit), convert_size_value(pz, size_unit),
                                       c_w, convert_size_value(row.get('PACKAGING SIZE - X (cm)', ''), size_unit),
                                       convert_size_value(row.get('PACKAGING SIZE - Y (cm)', ''), size_unit), convert_size_value(row.get('PACKAGING SIZE - Z (cm)', ''), size_unit)])

            out_df = pd.DataFrame(processed_data, columns=output_headers)
            st.dataframe(out_df)

            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                out_df.to_excel(writer, index=False, sheet_name='Sheet1')
                ws = writer.sheets['Sheet1']
                al_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
                al_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
                for c_idx, c_name in enumerate(output_headers, 1):
                    letter = ws.cell(row=1, column=c_idx).column_letter
                    if "Feature" in str(c_name): ws.column_dimensions[letter].width = 15
                    elif any(w in str(c_name) for w in ["PRICE", "SIZE", "WEIGHT", "PACKAGES"]):
                        m_data = max([len(str(ws.cell(row=r, column=c_idx).value)) for r in range(2, len(out_df)+2)] + [0])
                        ws.column_dimensions[letter].width = m_data + 5
                    else:
                        m_all = max([len(str(ws.cell(row=r, column=c_idx).value)) for r in range(1, len(out_df)+2)] + [0])
                        ws.column_dimensions[letter].width = min(m_all + 2, 40)
                    for r_idx in range(1, len(out_df) + 2):
                        cell = ws.cell(row=r_idx, column=c_idx)
                        cell.alignment = al_left if (r_idx > 1 and "Feature" in str(c_name)) else al_center
                        if r_idx > 1: ws.row_dimensions[r_idx].height = 15
                ws.row_dimensions[1].height = 45
            st.download_button("Download Processed Excel", output.getvalue(), output_filename, use_container_width=True)
        except Exception as e: st.error(f"Error: {e}")


# --- PAGE 2: PO TRACKING TOOL ---
elif page == "PO Tracking Tool":
    st.header("PO Tracking Tool")
    pdf_files = st.file_uploader("Upload PDF files", type="pdf", accept_multiple_files=True)
    if pdf_files:
        if st.button("Extract Data", use_container_width=True):
            with st.spinner("Analyzing PDF layers..."):
                results_df = process_pdfs_robust(pdf_files)
                if not results_df.empty:
                    st.dataframe(results_df)
                    current_date = datetime.now().strftime("%d-%m-%Y")
                    date_filename = f"{current_date}_Tracking_List.xlsx"
                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        results_df.to_excel(writer, index=False)
                    st.download_button(f"Download {date_filename}", output.getvalue(), date_filename, use_container_width=True)
                else: st.warning("No valid tracking numbers found.")


# --- PAGE 3: TITLE GENERATOR TOOL ---
elif page == "Title Generator Tool":
    st.header("📝 Ürün İsmi Oluşturucu (Title Generator)")
    st.markdown("Yüklediğiniz dosyadan, paket ölçüleri gibi gereksiz verileri ayıklayarak **Ana Ürün, Renk, Materyal ve Ölçü** bileşenlerinden oluşan temiz bir ürün ismi oluşturur.")
    
    with st.sidebar:
        st.subheader("Title Generator Settings")
        target_unit = st.radio("Oluşturulacak İsimdeki Ölçü Birimi:", ("inch", "cm"), index=0, horizontal=True)

    uploaded_file = st.file_uploader("Lütfen Veri Dosyanızı Yükleyin (Excel veya CSV)", type=["xlsx", "xls", "csv"])

    if uploaded_file is not None:
        try:
            if uploaded_file.name.endswith('.csv'):
                df = pd.read_csv(uploaded_file)
            else:
                df = pd.read_excel(uploaded_file)
                
            st.success("Dosya başarıyla yüklendi!")
            
            columns = df.columns.tolist()
            
            st.subheader("Sütun Eşleştirmeleri")
            col1, col2, col3 = st.columns(3)
            
            with col1:
                name_col = st.selectbox("Ana Ürün Adı / NAME Sütunu", options=columns)
            with col2:
                features_col = st.selectbox("Ek Özellikler / FEATURES Sütunu (Opsiyonel)", options=["Yok"] + columns)
            with col3:
                marketing_col = st.selectbox("Marketing Copy Sütunu (Opsiyonel)", options=["Yok"] + columns)

            if st.button("🚀 Ürün İsimlerini Oluştur", use_container_width=True):
                with st.spinner("Regex kuralları işleniyor ve isimler oluşturuluyor..."):
                    
                    f_col = features_col if features_col != "Yok" else None
                    m_col = marketing_col if marketing_col != "Yok" else None
                    
                    df['Generated_Title'] = df.apply(lambda row: generate_title(row, name_col, f_col, target_unit, m_col), axis=1)
                    
                    st.success("İşlem Tamamlandı! Aşağıdan önizlemeye bakabilirsiniz.")
                    
                    st.dataframe(df[[name_col, 'Generated_Title']].head(20))
                    
                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df.to_excel(writer, index=False, sheet_name='Sheet1')
                    
                    st.download_button(
                        label="📥 Yeni Excel Dosyasını İndir",
                        data=output.getvalue(),
                        file_name="Temiz_Urun_Isimleri.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )

        except Exception as e:
            st.error(f"Bir hata oluştu: {e}")
